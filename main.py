import copy
import os
import os.path as osp
import random
import time

import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
import wandb
from openpyxl import load_workbook
from torch.utils.data import DataLoader
from torchvision import transforms

from dataloader import LLP_dataset, ToTensor, categories
from mutual_loss import MutualLearningLoss
from nets.net_audiovisual import MMIL_Net, LabelSmoothingNCELoss
from option import build_args
from utils.eval_metrics import segment_level, event_level, print_overall_metric
from utils.write_excel import write_excel, create_empty_excel


def get_LLP_dataloader(args):
    train_dataset = LLP_dataset(label=args.label_train, audio_dir=args.audio_dir,
                                video_dir=args.video_dir, st_dir=args.st_dir,
                                transform=transforms.Compose([ToTensor()]),
                                a_smooth=args.a_smooth, v_smooth=args.v_smooth)
    val_dataset = LLP_dataset(label=args.label_val, audio_dir=args.audio_dir,
                              video_dir=args.video_dir, st_dir=args.st_dir,
                              transform=transforms.Compose([ToTensor()]))
    test_dataset = LLP_dataset(label=args.label_test, audio_dir=args.audio_dir,
                               video_dir=args.video_dir, st_dir=args.st_dir,
                               transform=transforms.Compose([ToTensor()]))

    train_loader = DataLoader(train_dataset, batch_size=args.batch_size,
                              shuffle=True, num_workers=5, pin_memory=True, sampler=None)
    val_loader = DataLoader(val_dataset, batch_size=1, shuffle=False,
                            num_workers=1, pin_memory=True, sampler=None)
    test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False,
                             num_workers=1, pin_memory=True, sampler=None)

    return train_loader, val_loader, test_loader


def set_random_seed(seed):
    torch.manual_seed(seed)
    torch.cuda.manual_seed(seed)
    random.seed(seed)
    np.random.seed(seed)


def get_random_state():
    state = {
        'torch_rng': torch.get_rng_state(),
        'cuda_rng': torch.cuda.get_rng_state(),
        'random_rng': random.getstate(),
        'numpy_rng': np.random.get_state()
    }
    return state


def train(args, model, train_loader, optimizer, criterion, epoch):
    print("----------------------------------------------------------")
    model.train()
    criterion2 = LabelSmoothingNCELoss(classes=10, smoothing=args.nce_smooth)

    noise_ratios = np.load(args.noise_ratio_file)
    noise_ratios_a_init = torch.from_numpy(noise_ratios['audio']).to('cuda')
    noise_ratios_v_init = torch.from_numpy(noise_ratios['visual']).to('cuda')
    noise_ratios_a = noise_ratios_a_init.clone()
    noise_ratios_v = noise_ratios_v_init.clone()

    iters_per_epoch = len(train_loader)

    for batch_idx, sample in enumerate(train_loader):
        audio, video, video_st, target = sample['audio'].to('cuda'), \
                                         sample['video_s'].to('cuda'), \
                                         sample['video_st'].to('cuda'), \
                                         sample['label'].type(torch.FloatTensor).to('cuda')
        Pa, Pv = sample['Pa'].type(torch.FloatTensor).to('cuda'), sample['Pv'].type(torch.FloatTensor).to('cuda')
        batch = len(audio)

        if args.warm_up_epoch is not None:
            noise_ratios_a = \
                torch.min(
                    torch.cat(
                        (noise_ratios_a.reshape(1, -1),
                         noise_ratios_a.reshape(1, -1) *
                         ((epoch - 1) * iters_per_epoch + batch_idx) / (args.warm_up_epoch * iters_per_epoch)),
                        dim=0),
                    dim=0)[0]
            noise_ratios_v = \
                torch.min(
                    torch.cat(
                        (noise_ratios_v.reshape(1, -1),
                         noise_ratios_v.reshape(1, -1) *
                         ((epoch - 1) * iters_per_epoch + batch_idx) / (args.warm_up_epoch * iters_per_epoch)),
                        dim=0),
                    dim=0)[0]

        with torch.no_grad():
            output, a_prob, v_prob, frame_prob = model(audio, video, video_st, with_ca=False)[:4]

            a_prob = torch.clamp(a_prob, min=args.clamp, max=1 - args.clamp)
            v_prob = torch.clamp(v_prob, min=args.clamp, max=1 - args.clamp)

            tmp_loss_a = nn.BCELoss(reduction='none')(a_prob, Pa)
            tmp_loss_v = nn.BCELoss(reduction='none')(v_prob, Pv)
            _, sort_index_a = torch.sort(tmp_loss_a, dim=0)
            _, sort_index_v = torch.sort(tmp_loss_v, dim=0)

            pos_index_a = Pa > 0.5
            pos_index_v = Pv > 0.5

            for i in range(25):
                pos_num_a = sum(pos_index_a[:, i].type(torch.IntTensor))
                pos_num_v = sum(pos_index_v[:, i].type(torch.IntTensor))
                numbers_a = torch.mul(noise_ratios_a[i], pos_num_a).type(torch.IntTensor)
                numbers_v = torch.mul(noise_ratios_v[i], pos_num_v).type(torch.IntTensor)
                # remove noise labels for visual
                mask_a = torch.zeros(batch).to('cuda')
                mask_v = torch.zeros(batch).to('cuda')
                if numbers_v > 0:
                    mask_a[sort_index_a[pos_index_v[sort_index_a[:, i], i], i][:numbers_v]] = 1
                    mask_v[sort_index_v[pos_index_v[sort_index_v[:, i], i], i][-numbers_v:]] = 1
                mask = torch.nonzero(torch.mul(mask_a, mask_v)).squeeze(-1).type(torch.LongTensor)
                Pv[mask, i] = 0

                # remove noise labels for audio
                mask_a = torch.zeros(batch).to('cuda')
                mask_v = torch.zeros(batch).to('cuda')
                if numbers_a > 0:
                    mask_a[sort_index_a[pos_index_a[sort_index_a[:, i], i], i][-numbers_a:]] = 1
                    mask_v[sort_index_v[pos_index_a[sort_index_v[:, i], i], i][:numbers_a]] = 1
                mask = torch.nonzero(torch.mul(mask_a, mask_v)).squeeze(-1).type(torch.LongTensor)
                Pa[mask, i] = 0

        optimizer.zero_grad()
        output, a_prob, v_prob, frame_prob, sims_after, mask_after, global_uct, a_uct, v_uct, frame_uct = \
            model(audio, video, video_st, with_ca=True)

        output = torch.clamp(output, min=args.clamp, max=1 - args.clamp)
        a_prob = torch.clamp(a_prob, min=args.clamp, max=1 - args.clamp)
        v_prob = torch.clamp(v_prob, min=args.clamp, max=1 - args.clamp)

        loss1 = criterion(a_prob, Pa)
        loss2 = criterion(v_prob, Pv)
        loss3 = criterion(output, target)

        loss4 = criterion2(sims_after, mask_after)

        criterion_mutual = MutualLearningLoss(eta=args.mutual_eta)
        loss_mutual = criterion_mutual(target, output, a_prob, v_prob, a_uct, v_uct, global_uct, batch_idx)

        loss = loss1 * args.audio_weight + loss2 * args.visual_weight + \
               loss3 * args.video_weight + loss4 * args.nce_weight + \
               loss_mutual * args.mutual_weight

        loss.backward()
        optimizer.step()

        if batch_idx % args.log_interval == 0:
            print(f'Train Epoch: {epoch}\n'
                  f'L1: {loss1.item():.4f}\tL2: {loss2.item():.4f}\tL3: {loss3.item():.4f}\tL4: {loss4.item():.4f}\t'
                  f'L5: {loss_mutual.item():.4f}')

        if not args.without_wandb:
            loss_dict = {
                'loss': loss,
                'loss_audio': loss1,
                'loss_visual': loss2,
                'loss_global': loss3,
                'loss_nce': loss4,
                'loss_mutual': loss_mutual,
            }
            wandb.log(loss_dict, step=(epoch - 1) * len(train_loader) + batch_idx + 1)


def eval(args, model, val_loader, set, v_thres=0.4, target_class=None):
    model.eval()
    # load annotations
    df = pd.read_csv(set, header=0, sep='\t')
    df_a = pd.read_csv(args.eval_audio, header=0, sep='\t')
    df_v = pd.read_csv(args.eval_video, header=0, sep='\t')

    id_to_idx = {id: index for index, id in enumerate(categories)}
    F_seg_a = []
    F_seg_v = []
    F_seg = []
    F_seg_av = []
    F_event_a = []
    F_event_v = []
    F_event = []
    F_event_av = []

    with torch.no_grad():
        for batch_idx, sample in enumerate(val_loader):
            audio, video, video_st, target = sample['audio'].to('cuda'), \
                                             sample['video_s'].to('cuda'), \
                                             sample['video_st'].to('cuda'), \
                                             sample['label'].to('cuda')

            a_prob1, v_prob1, frame_prob1 = model(audio, video, video_st, with_ca=args.with_ca)[1:4]
            a_prob2, v_prob2, frame_prob2 = model(audio, video, video_st, with_ca=False)[1:4]

            a_prob = a_prob1 * args.fuse_ratio + a_prob2 * (1 - args.fuse_ratio)
            v_prob = v_prob1 * args.fuse_ratio + v_prob2 * (1 - args.fuse_ratio)
            frame_prob = frame_prob1 * args.fuse_ratio + frame_prob2 * (1 - args.fuse_ratio)

            v_thres_list = np.full((1, 25), 0.45)

            if args.mode == 'select_thresholds':
                v_thres_list[0, target_class] = v_thres

            if args.mode == 'test':
                excel_path = osp.join(args.model_save_dir, args.group_name, args.exp_name) + '_early.xlsx'
                wb = load_workbook(excel_path)
                ws = wb['Thres']
                for idx in range(25):
                    v_thres_list[0, idx] = ws['A'][idx].value

            oa = (a_prob.cpu().detach().numpy() >= 0.45).astype(np.int_)
            ov = (v_prob.cpu().detach().numpy() >= v_thres_list).astype(np.int_)

            Pa = frame_prob[0, :, 0, :].cpu().detach().numpy()
            Pv = frame_prob[0, :, 1, :].cpu().detach().numpy()

            # filter out false positive events with predicted weak labels
            Pa = (Pa >= 0.25).astype(np.int_) * np.repeat(oa, repeats=10, axis=0)
            Pv = (Pv >= 0.25).astype(np.int_) * np.repeat(ov, repeats=10, axis=0)

            # extract audio GT labels
            GT_a = np.zeros((25, 10))
            GT_v = np.zeros((25, 10))

            df_vid_a = df_a.loc[df_a['filename'] == df.loc[batch_idx, :][0]]
            filenames = df_vid_a["filename"]
            events = df_vid_a["event_labels"]
            onsets = df_vid_a["onset"]
            offsets = df_vid_a["offset"]
            num = len(filenames)
            if num > 0:
                for i in range(num):
                    x1 = int(onsets[df_vid_a.index[i]])
                    x2 = int(offsets[df_vid_a.index[i]])
                    event = events[df_vid_a.index[i]]
                    idx = id_to_idx[event]
                    GT_a[idx, x1:x2] = 1

            # extract visual GT labels
            df_vid_v = df_v.loc[df_v['filename'] == df.loc[batch_idx, :][0]]
            filenames = df_vid_v["filename"]
            events = df_vid_v["event_labels"]
            onsets = df_vid_v["onset"]
            offsets = df_vid_v["offset"]
            num = len(filenames)
            if num > 0:
                for i in range(num):
                    x1 = int(onsets[df_vid_v.index[i]])
                    x2 = int(offsets[df_vid_v.index[i]])
                    event = events[df_vid_v.index[i]]
                    idx = id_to_idx[event]
                    GT_v[idx, x1:x2] = 1

            GT_av = GT_a * GT_v

            # obtain prediction matrices
            SO_a = np.transpose(Pa)
            SO_v = np.transpose(Pv)
            SO_av = SO_a * SO_v

            # segment-level F1 scores
            f_a, f_v, f, f_av = segment_level(SO_a, SO_v, SO_av, GT_a, GT_v, GT_av)
            F_seg_a.append(f_a)
            F_seg_v.append(f_v)
            F_seg.append(f)
            F_seg_av.append(f_av)

            # event-level F1 scores
            f_a, f_v, f, f_av = event_level(SO_a, SO_v, SO_av, GT_a, GT_v, GT_av)
            F_event_a.append(f_a)
            F_event_v.append(f_v)
            F_event.append(f)
            F_event_av.append(f_av)

    if args.mode == 'select_thresholds':
        return print_overall_metric(F_seg_a, F_seg_v, F_seg, F_seg_av, F_event_a, F_event_v, F_event, F_event_av, verbose=False)
    else:
        return print_overall_metric(F_seg_a, F_seg_v, F_seg, F_seg_av, F_event_a, F_event_v, F_event, F_event_av, verbose=True)


def main():
    args = build_args()

    os.environ['CUDA_VISIBLE_DEVICES'] = f'{args.gpu}'

    if not args.without_wandb:
        wandb_name = time.asctime()[:-4] + args.group_name + " " + args.exp_name
        wandb.init(name=wandb_name, config=vars(args), group=args.group_name, project=f"CVPR23_AVVP")

    # print parameters
    if not args.mode == 'select_thresholds':
        print('----------------args-----------------')
        for k in list(vars(args).keys()):
            print('%s: %s' % (k, vars(args)[k]))
        print('----------------args-----------------')
        cur = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        print(f'current time: {cur}')

    set_random_seed(args.seed)
    if not args.not_save:
        os.makedirs(osp.join(args.model_save_dir, args.group_name), exist_ok=True)

    model = MMIL_Net(args.num_layers, args.temperature).to('cuda')

    start = time.time()

    if args.mode == 'train':
        args.with_ca = True

        train_loader, val_loader, test_loader = get_LLP_dataloader(args)

        optimizer = optim.Adam(model.parameters(), lr=args.lr, weight_decay=args.weight_decay)
        scheduler = optim.lr_scheduler.StepLR(optimizer, step_size=args.lr_step_size, gamma=args.lr_gamma)

        criterion = nn.BCELoss()

        best_F = 0
        best_model = None
        best_epoch = 0

        for epoch in range(1, args.epochs + 1):
            train(args, model, train_loader, optimizer, criterion, epoch=epoch)
            scheduler.step()

            print(f"Test Epoch: {epoch}")
            audio_seg, visual_seg, av_seg, avg_type_seg, avg_event_seg, \
            audio_eve, visual_eve, av_eve, avg_type_eve, avg_event_eve, avg_all \
                = eval(args, model, val_loader, args.label_val)

            select_metric = avg_all

            if select_metric >= best_F:
                best_F = select_metric
                best_model = copy.deepcopy(model)
                best_epoch = epoch
            print(f'Best_epoch: {best_epoch}, Best_avg: \033[34m{best_F:.2f}\033[0m.')
            if not args.without_wandb:
                log_dict = {
                    "audio_seg": audio_seg,
                    "visual_seg": visual_seg,
                    "av_seg": av_seg,
                    "avg_type_seg": avg_type_seg,
                    "avg_event_seg": avg_event_seg,
                    "audio_eve": audio_eve,
                    "visual_eve": visual_eve,
                    "av_eve": av_eve,
                    "avg_type_eve": avg_type_eve,
                    "avg_event_eve": avg_event_eve,
                }
                wandb.log(log_dict, step=epoch * len(train_loader))

            if epoch == args.early_save_epoch and not args.not_save:
                state_dict = get_random_state()
                state_dict['model'] = model.state_dict()
                state_dict['optimizer'] = optimizer.state_dict()
                state_dict['scheduler'] = scheduler.state_dict()
                state_dict['epochs'] = args.epochs
                torch.save(state_dict, osp.join(args.model_save_dir, args.group_name, args.exp_name + '_early.pt'))

        optimizer.zero_grad()
        model = best_model
        if not args.not_save:
            state_dict = get_random_state()
            state_dict['model'] = model.state_dict()
            state_dict['optimizer'] = optimizer.state_dict()
            state_dict['scheduler'] = scheduler.state_dict()
            state_dict['epochs'] = args.epochs
            torch.save(state_dict, osp.join(args.model_save_dir, args.group_name, args.exp_name + '.pt'))
        print("----------------------------------------------------------")
        print(f"Test the best epoch {best_epoch} model:")
        eval(args, model, test_loader, args.label_test)

    elif args.mode == 'test':
        dataset = args.label_test
        args.with_ca = True if args.mode == 'test' else False
        test_dataset = LLP_dataset(label=dataset,
                                   audio_dir=args.audio_dir, video_dir=args.video_dir, st_dir=args.st_dir,
                                   transform=transforms.Compose([ToTensor()]))
        test_loader = DataLoader(test_dataset, batch_size=1, shuffle=False, num_workers=1, pin_memory=True)
        resume = torch.load(osp.join(args.model_save_dir, args.group_name, args.exp_name + '.pt'))
        model.load_state_dict(resume['model'])
        eval(args, model, test_loader, dataset)

    elif args.mode == 'select_thresholds':
        args.with_ca = True
        _, val_loader, test_loader = get_LLP_dataloader(args)
        resume = torch.load(osp.join(args.model_save_dir, args.group_name, args.exp_name + '_early.pt'))
        model.load_state_dict(resume['model'])
        thres_candi = [0.25, 0.3, 0.35, 0.4, 0.45, 0.5, 0.55, 0.6, 0.65, 0.7]

        excel_path = osp.join(args.model_save_dir, args.group_name, args.exp_name) + '_early.xlsx'
        if not os.path.exists(excel_path):
            create_empty_excel(excel_path)

        for target_class in range(args.start_class, args.start_class + 5):
            best_select_metric = 0
            best_v_thres = 0
            for i, v_thres in enumerate(thres_candi):
                audio_seg, visual_seg, av_seg, avg_type_seg, avg_event_seg, \
                audio_eve, visual_eve, av_eve, avg_type_eve, avg_event_eve, avg_all \
                    = eval(args, model, val_loader, args.label_val, v_thres=v_thres, target_class=target_class)

                select_metric = visual_eve

                if select_metric >= best_select_metric:
                    best_select_metric = select_metric
                    best_v_thres = v_thres

            write_excel(excel_path, target_class, best_v_thres)
            print(f'class {target_class} done')

    if not args.mode == 'select_thresholds':
        end = time.time()
        print(f'duration time {(end - start) / 60:.2f} mins.')
        cur = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(time.time()))
        print(f'current time: {cur}')


if __name__ == '__main__':
    main()
