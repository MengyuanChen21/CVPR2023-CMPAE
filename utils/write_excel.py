from openpyxl import Workbook, load_workbook


def create_empty_excel(path):
    wb = Workbook()
    wb.create_sheet('Thres', 0)
    wb.save(path)
    wb.close()


def write_excel(excel_path, target_class, v_thres):
    wb = load_workbook(excel_path)
    ws = wb['Thres']
    ws.cell(row=target_class + 1, column=1).value = v_thres
    wb.save(excel_path)
    wb.close()
    return 0
